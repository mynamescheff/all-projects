import openpyxl
from itertools import combinations

def find_combination(transactions, target_sum):
    # Check all possible combinations of transactions to find a subset that sums to target_sum
    total_combinations = sum(1 for i in range(1, len(transactions) + 1) for _ in combinations(transactions, i))
    print(f"Total combinations to check: {total_combinations}")
    count = 0
    for i in range(1, len(transactions) + 1):
        for combo in combinations(transactions, i):
            count += 1
            if count % 1000 == 0:
                print(f"Checked {count}/{total_combinations} combinations")
            if abs(sum(combo) - target_sum) < 1e-9:  # Allowing a small margin for floating point precision
                return combo
    return None

def find_transactions_for_all_people(file_path, target_sum):
    # Load the workbook and select the active sheet
    print("Loading workbook...")
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Collect all transactions for each person
    transactions_by_person = {}
    print("Collecting transactions by person...")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        person, transaction = row
        if transaction is not None:  # Ensure the transaction is not None
            try:
                transaction = float(transaction)  # Ensure the transaction is a float
                if transaction <= target_sum:  # Disregard transactions higher than the target value
                    if person in transactions_by_person:
                        transactions_by_person[person].append(transaction)
                    else:
                        transactions_by_person[person] = [transaction]
            except ValueError:
                # Skip the row if transaction is not a valid number
                continue
    print(f"Collected transactions for {len(transactions_by_person)} unique people.")

    # Find combinations of transactions for each person that sum up to the target sum
    results = {}
    for person, transactions in transactions_by_person.items():
        print(f"Finding combinations for {person} with {len(transactions)} transactions...")
        result = find_combination(transactions, target_sum)
        results[person] = result
        if result:
            print(f"Found a combination for {person}: {result}")
        else:
            print(f"No combination found for {person}.")

    return results

# Example usage
file_path = 'C:\\IT project3\\Book4.xlsx'  # Path to your Excel file
target_sum = 11.69  # Target sum to find

print("Starting transaction search...")
results = find_transactions_for_all_people(file_path, target_sum)
for person, result in results.items():
    if result:
        print(f"The transactions for {person} that sum up to {target_sum} are: {result}")
    else:
        print(f"No combination of transactions for {person} sums up to {target_sum}.")
