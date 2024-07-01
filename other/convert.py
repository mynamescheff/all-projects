import openpyxl

def find_and_write_missing_values(excel_file, sheet_name):
    # Load the workbook and the specified sheet
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Extract values from columns A and C
    values_A = [cell.value for cell in ws['A'] if cell.value is not None]
    values_C = [cell.value for cell in ws['C'] if cell.value is not None]
    
    # Create a dictionary to count occurrences of each value in column A and C
    count_A = {}
    count_C = {}
    
    for value in values_A:
        if value in count_A:
            count_A[value] += 1
        else:
            count_A[value] = 1
            
    for value in values_C:
        if value in count_C:
            count_C[value] += 1
        else:
            count_C[value] = 1
            
    # Identify missing values
    missing_values = []
    
    for value in count_C:
        count_in_A = count_A.get(value, 0)
        count_in_C = count_C[value]
        if count_in_A < count_in_C:
            missing_count = count_in_C - count_in_A
            missing_values.extend([value] * missing_count)
    
    # Write missing values to column E
    for i, value in enumerate(missing_values, start=1):
        ws[f'E{i}'] = value
    
    # Save the workbook
    wb.save(excel_file)

# Example usage
excel_file = 'your_excel_file.xlsx'
sheet_name = 'Sheet1'  # Change this to your sheet name

find_and_write_missing_values(excel_file, sheet_name)

print("Missing values have been written to column E.")
