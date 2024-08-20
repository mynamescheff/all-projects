import openpyxl
import csv
import logging
import os
import subprocess
from datetime import datetime
import re

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def determine_currency_bank_acc(currency, comparison_currency):
    if comparison_currency == "ACC LTD":
        return "GBP bank acc"
    elif comparison_currency == "ACC LTD USD":
        return "USD bank acc"
    elif comparison_currency == "ACC LTD EUR":
        return "EUR bank acc"
    elif comparison_currency == "ACC LTD CHF":
        return "CHF bank acc"
    else:
        if currency in ["EUR", "USD", "CHF"]:
            return f"{currency} bank acc"
        else:
            return "GBP bank acc"

def clean_column_b_value(value):
    return ''.join([char for char in value if char.isalpha() or char in ['-', '–', ' ']])

def split_b_value(b_value):
    parts = []
    while len(b_value) > 35:
        split_point = 35
        while split_point > 0 and not b_value[split_point].isalnum():
            split_point -= 1
        if split_point == 0:
            split_point = 35
        parts.append(b_value[:split_point].rstrip())
        b_value = b_value[split_point:].lstrip()
    parts.append(b_value)
    return parts

def clean_currency(value):
    if value:
        return value.replace('USD', '').replace('EUR', '')
    return value

def remove_duplicates(b_value):
    b_value = b_value.replace('\n', '').replace('\r', '')
    words = b_value.split()
    seen = set()
    result = []
    for word in words:
        if word not in seen:
            seen.add(word)
            result.append(word)
    return ' '.join(result)

def adjust_length(b_value, c_value):
    total_length = len(b_value) + len(c_value)
    if total_length <= 140:
        return b_value, c_value

    b_words = b_value.split()
    while len(' '.join(b_words)) + len(c_value) > 140:
        if len(b_words) > 1:
            b_words.pop()
        else:
            break
    return ' '.join(b_words), c_value

def check_special_transfer(comparison_sheet, comparison_row):
    transfer_type = comparison_sheet[f'F{comparison_row}'].value
    if transfer_type in ["UK Faster/Next Day Payment", "SEPA Credit Transfer"]:
        return transfer_type
    return None

try:
    output_dir = 'C:\\IT project3\\utils'
    log_dir = os.path.join(output_dir, 'import_logs')

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    comparison_path = os.path.join(output_dir, 'comparison_file.xlsx')
    logging.info(f'Loading comparison file from {comparison_path}')
    comparison_workbook = openpyxl.load_workbook(comparison_path)
    comparison_sheet = comparison_workbook.active
    comparison_dict = {
        clean_currency(comparison_sheet[f'C{row}'].value): row
        for row in range(2, comparison_sheet.max_row + 1)
    }
    logging.info('Comparison file loaded successfully')

    workbook_path = os.path.join(output_dir, 'combined_file.xlsx')
    logging.info(f'Loading workbook from {workbook_path}')
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook['Transposed']
    logging.info('Workbook loaded successfully')

    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'I']
    logging.debug(f'Columns to extract: {columns}')

    csv_path = os.path.join(output_dir, 'output.csv')
    logging.info(f'Opening CSV file at {csv_path} for writing')

    with open(csv_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        log_entries = []
        row_number = 1

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            row_number += 1

            if row[0].value is None:
                logging.info('Empty cell found in first column, stopping iteration')
                break

            # Clean B value to retain only letters and '-' or '–'
            col_b_value = clean_column_b_value(row[sheet['B1'].column - 1].value or '')
            col_c_value = (row[sheet['C1'].column - 1].value or '').replace('\n', '').replace('\r', '')
            col_e_value = (row[sheet['E1'].column - 1].value or '').replace('\n', '').replace('\r', '')
            col_f_value = clean_currency((row[sheet['F1'].column - 1].value or '').replace('\n', '').replace('\r', ''))
            col_g_value = clean_currency((row[sheet['G1'].column - 1].value or '').replace('\n', '').replace('\r', ''))

            # Deduplicate F and G values
            match_value = None
            comparison_row = None
            matched_column = None

            if col_f_value in comparison_dict:
                match_value = col_f_value
                comparison_row = comparison_dict[col_f_value]
                matched_column = 'F'
            if col_g_value in comparison_dict and col_g_value != col_f_value:
                match_value = col_g_value
                comparison_row = comparison_dict[col_g_value]
                matched_column = 'G'

            if match_value is None:
                reason = f"Row {row_number} (Name: '{col_b_value}'): Neither F nor G matched with comparison file."
                log_entries.append(reason)
                continue

            special_transfer = check_special_transfer(comparison_sheet, comparison_row)
            if special_transfer:
                reason = f"Row {row_number} (Name: '{col_b_value}'): Value '{match_value}' skipped because it's a special transfer type ({special_transfer})."
                log_entries.append(reason)
                continue

            # Move E value before the matched value (either F or G)
            trimmed_b_16 = col_b_value[:16]
            b_parts = split_b_value(col_b_value)
            if len(b_parts) > 3:
                b_value = remove_duplicates(col_b_value)
                b_parts = split_b_value(b_value)

            while len(b_parts) < 3:
                b_parts.append('')

            b_parts.append(col_c_value)

            # Include the matched value only (either F or G)
            other_values = [col_e_value, match_value]

            other_values.extend([row[sheet[column + '1'].column - 1].value for column in columns[4:5]])

            comparison_currency_value = comparison_sheet[f'E{comparison_row}'].value

            currency_bank_acc = determine_currency_bank_acc(col_e_value, comparison_currency_value)

            row_data = [trimmed_b_16] + b_parts + other_values + [currency_bank_acc]
            logging.debug(f'Writing row data: {row_data}')
            writer.writerow(row_data)

    if log_entries:
        log_path = os.path.join(log_dir, f'log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt')
        with open(log_path, 'w') as log_file:
            log_file.write(f"Log generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            log_file.write("\n".join(log_entries))
        logging.info(f'Log file created at {log_path}')

        # Automatically open the log file
        try:
            subprocess.Popen(['notepad.exe', log_path])
        except FileNotFoundError:
            logging.warning("Notepad not found. Trying to open with the default editor.")
            try:
                os.startfile(log_path)
            except Exception as e:
                logging.error(f"Could not open log file: {e}", exc_info=True)

    logging.info('Data successfully written to CSV file')

except Exception as e:
    logging.error(f'An error occurred: {e}', exc_info=True)
