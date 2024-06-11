#create a function to concatenate all sheets in an excel file to a single sheet with all the contents of the sheets

import os
import openpyxl

def concatenate_sheets(file_path):
    #get absolute directory path
    path = os.path.abspath(os.path.dirname(__file__))
    print(path)

    #load the excel file
    wb = openpyxl.load_workbook(os.path.join(path, file_path))
    print(wb.sheetnames)

    #create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    #loop through each sheet in the workbook
    for sheet in wb.sheetnames:
        current_sheet = wb[sheet]
        for row in current_sheet.iter_rows():
            new_sheet.append([cell.value for cell in row])

    new_wb.save(os.path.join(path, 'concatenated.xlsx'))

concatenate_sheets('data.xlsx')
