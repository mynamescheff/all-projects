import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelDuplicateChecker:
    def __init__(self, directory, filename):
        self.directory = directory
        self.filename = filename
        self.filepath = os.path.join(directory, filename)
        self.df = pd.read_csv(self.filepath)

    def check_duplicates_in_column_d(self):
        duplicates = self.df[self.df.duplicated(subset=['D'], keep=False)]
        return duplicates

    def load_excel_file(self, excel_filename):
        excel_filepath = os.path.join(self.directory, excel_filename)
        wb = load_workbook(excel_filepath)
        return wb

    def process_duplicates(self, excel_filename):
        duplicates = self.check_duplicates_in_column_d()
        if duplicates.empty:
            print("No duplicates found in column D.")
            return
        
        print("Duplicates found:")
        print(duplicates)
        
        wb = self.load_excel_file(excel_filename)
        ws = wb.active

        # Assuming we want to highlight duplicates in the Excel file
        for row in ws.iter_rows(min_row=2, max_col=4):
            cell = row[3]  # Column D is the 4th column (index 3)
            if cell.value in duplicates['D'].values:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        wb.save(excel_filename)
        print(f"Duplicates highlighted in {excel_filename}")

# Example usage:
# checker = ExcelDuplicateChecker('/path/to/directory', 'file.csv')
# checker.process_duplicates('file.xlsx')
