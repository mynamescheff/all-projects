from openpyxl import load_workbook
from datetime import date
import os

class ExcelComparator:
    def __init__(self, combined_file, sprawdzacz_file, output_path):
        self.combined_file = combined_file
        self.sprawdzacz_file = sprawdzacz_file
        self.output_path = output_path

    def compare_and_append(self):
        try:
            # Load the workbooks
            combined_workbook = load_workbook(filename=self.combined_file)
            combined_sheet = combined_workbook['Transposed']
            sprawdzacz_workbook = load_workbook(filename=self.sprawdzacz_file)
            sprawdzacz_sheet = sprawdzacz_workbook['name_acc']

            # Initialize lists for non-matching values and their corresponding filenames
            non_matching_values = []
            filenames = []

            # Iterate through rows and check for matches
            for row in range(2, combined_sheet.max_row + 1):  # Skip header
                case_value = combined_sheet[f'A{row}'].value
                f_value = str(combined_sheet[f'F{row}'].value).replace(" ", "").replace("-", "") if combined_sheet[f'F{row}'].value else ""
                h_value = str(combined_sheet[f'G{row}'].value).replace(" ", "").replace("-", "") if combined_sheet[f'G{row}'].value else ""

                found_match = any(f_value == str(row_val[2]) or h_value == str(row_val[2]) for row_val in sprawdzacz_sheet.iter_rows(values_only=True))

                if not found_match:
                    non_matching_values.append((case_value, f_value, h_value))
                    filenames.append((self.combined_file, self.sprawdzacz_file))

            # Write to the output file
            file_path = os.path.join(self.output_path, 'mismatch_list.txt')
            with open(file_path, 'w', encoding="utf-8") as file:
                current_date = date.today()
                for value in non_matching_values:
                    modified_value = value[0].replace("\xa0", "").strip()
                    file.write(f"{modified_value}: {value[1]}, {value[2]} ({current_date})\n")

            return non_matching_values, f"Non-matching values appended to {file_path} successfully."
        except Exception as e:
            return [], str(e)

# This setup allows for the functions to be directly called from a GUI handler
